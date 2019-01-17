#!/usr/bin/env python
# _*_ UTF-8 _*_
import openpyxl
import os
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.chart import BarChart, Series, Reference, BarChart3D
from openpyxl.styles import Color, Font, Alignment
from openpyxl.styles import Border, Side
from openpyxl.styles.colors import BLUE, RED, GREEN, YELLOW
from openpyxl.styles import PatternFill

def create_excel(context_list,title_list,sheet_list):
    # 1、创建excel文件
    f = openpyxl.Workbook()
    for i in range(int(sheet_list[0])):
        f.create_sheet(title=sheet_list[i+1], index=i)
    # 2、创建sheet格式
    # sheet:
    sheet = f.worksheets[2]
    for i in range(0,5):
        # 行合并
        sheet.merge_cells(start_row=16*i+3, start_column=2, end_row=16*i+18, end_column=2)
        sheet.merge_cells(start_row=16*i+3, start_column=3, end_row=16*i+6, end_column=3)
        sheet.merge_cells(start_row=16*i+7, start_column=3, end_row=16*i+10, end_column=3)
        sheet.merge_cells(start_row=16*i+11, start_column=3, end_row=16*i+14, end_column=3)
        sheet.merge_cells(start_row=16*i+15, start_column=3, end_row=16*i+18, end_column=3)
        sheet.merge_cells(start_row=16*i+3, start_column=4, end_row=16*i+6, end_column=4)
        sheet.merge_cells(start_row=16*i+7, start_column=4, end_row=16*i+10, end_column=4)
        sheet.merge_cells(start_row=16*i+11, start_column=4, end_row=16*i+14, end_column=4)
        sheet.merge_cells(start_row=16*i+15, start_column=4, end_row=16*i+18, end_column=4)
    sheet.merge_cells(start_row=83, start_column=2, end_row=83, end_column=4)
    sheet.merge_cells(start_row=84, start_column=2, end_row=84, end_column=4)
    sheet.merge_cells(start_row=85, start_column=2, end_row=85, end_column=4)

    # 设置单元格字体
    # font = Font(name=u'宋体', size=14, color=BLUE, bold=True)
    # sheet.cell(row=1, column=i+1).font = font
    align = Alignment(horizontal='center', vertical='center')
    fill = PatternFill(start_color='CC99CC', end_color='CC99CC', fill_type='solid')
    sheet.freeze_panes = 'F3'
    border = Border(left=Side(style='medium', color='FF000000'), right=Side(style='medium', color='FF000000'),
                    top=Side(style='medium', color='FF000000'), bottom=Side(style='medium', color='FF000000'),
                    diagonal=Side(style='medium', color='FF000000'), diagonal_direction=0,
                    outline=Side(style='medium', color='FF000000'), vertical=Side(style='medium', color='FF000000'),
                    horizontal=Side(style='medium', color='FF000000'))

    for i in range(2,86):
        for j in range(2,37):
            sheet.cell(row=i, column=j).border = border

    # 3、填写内容
    for i in range(len(context_list)):
        sheet.cell(row=2, column=i+2).alignment = align
        sheet.cell(row=2, column=i+2, value=10).fill = fill
        sheet.cell(row=2, column=i+2).value = context_list[i]

    B_list = ['1Q', '2Q', '3Q', '4Q', '全年']
    # for i in range(len(B_list)):
    #     sheet.cell(row=i+3, column=2).alignment = align
    #     sheet.cell(row=i+3, column=2).value = B_list[i]
    sheet.cell(row=3, column=2).alignment = align
    sheet.cell(row=3, column=2).value = B_list[0]
    sheet.cell(row=19, column=2).alignment = align
    sheet.cell(row=19, column=2).value = B_list[1]
    sheet.cell(row=35, column=2).alignment = align
    sheet.cell(row=35, column=2).value = B_list[2]
    sheet.cell(row=51, column=2).alignment = align
    sheet.cell(row=51, column=2).value = B_list[3]
    sheet.cell(row=67, column=2).alignment = align
    sheet.cell(row=67, column=2).value = B_list[4]

    C_list = ['签约1个月', '签约半年', '交付一周', '入住期']
    for j in range(len(C_list)):
        for l in range(0,4):
            sheet.cell(row=j + (4*l+3), column=3).alignment = align
            sheet.cell(row=j + (4*l+3), column=3).value = C_list[j]

    D_list = ["2018.12-2019.2签约",
            "2018.7-2018.9签约",
            "2018.12.5-2019.3.15交付",
            "\\",
            "2019.3-2019.5签约",
            "2018.10-2018.12签约",
            "2019.3.16-2019.6.15交付",
            "\\",
            "2019.6-2019.8签约",
            "2019.1-2019.3签约",
            "2019.6.16-2019.9.15交付",
            "\\",
            "2019.9-2019.11签约",
            "2019.4-2019.6签约",
            "2019.9.16-2019.12.10交付",
            "\\",
            "2018.12-2019.11签约",
            "2018.7-2019.6签约",
            "2018.12.5-2019.12.10交付",
            "\\"]
    for i in range(len(D_list)):
        sheet.cell(row=i+3, column=4).alignment = align
        sheet.cell(row=i+3, column=4).value = D_list[i]

    # 4、保存文件
    curPath = os.getcwd()
    tempPath = 'file_target'
    targetPath = curPath + os.path.sep + tempPath
    if not os.path.exists(targetPath):
        os.makedirs(targetPath)
    for i in range(len(title_list)):
        f.save(targetPath+"/"+title_list[i]+".xlsx")

if __name__ == '__main__':
    context_list = ["季度",
                    "调研期别",
                    "调研对象",
                    "项目",
                    "合计",
                    "项目1",
                    "项目2",
                    "项目3",
                    "项目4",
                    "项目5",
                    "项目6",
                    "项目7",
                    "项目8",
                    "项目9",
                    "项目10",
                    "项目11",
                    "项目12",
                    "项目13",
                    "项目14",
                    "项目15",
                    "项目16",
                    "项目17",
                    "项目18",
                    "项目19",
                    "项目20",
                    "项目21",
                    "项目22",
                    "项目23",
                    "项目24",
                    "项目25",
                    "项目26",
                    "项目27",
                    "项目28",
                    "项目29",
                    "项目30",]


    title_list = ['2018满意度样本测算及管控（全周期）']
    sheet_list = ['4','说明','区域合计','区域项目','sheet3']
    create_excel(context_list,title_list,sheet_list)







