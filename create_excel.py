#!/usr/bin/env python
# _*_ UTF-8 _*_
import openpyxl
import os
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.chart import BarChart, Series, Reference, BarChart3D
from openpyxl.styles import Color, Font, Alignment
from openpyxl.styles import Border, Side
from openpyxl.styles.colors import BLUE, RED, GREEN, YELLOW
from openpyxl.styles import PatternFill

def create_excel(context_list,title_list,sheet_list):
    # 1、创建excel文件
    f = openpyxl.Workbook()
    for i in range(int(sheet_list[0])):
        f.create_sheet(title=sheet_list[i+1], index=i)
    # 2、创建sheet格式
    # sheet:
    sheet = f.worksheets[2]
    for i in range(0,5):
        # 行合并
        sheet.merge_cells(start_row=16*i+3, start_column=2, end_row=16*i+18, end_column=2)
        sheet.merge_cells(start_row=16*i+3, start_column=3, end_row=16*i+6, end_column=3)
        sheet.merge_cells(start_row=16*i+7, start_column=3, end_row=16*i+10, end_column=3)
        sheet.merge_cells(start_row=16*i+11, start_column=3, end_row=16*i+14, end_column=3)
        sheet.merge_cells(start_row=16*i+15, start_column=3, end_row=16*i+18, end_column=3)
        sheet.merge_cells(start_row=16*i+3, start_column=4, end_row=16*i+6, end_column=4)
        sheet.merge_cells(start_row=16*i+7, start_column=4, end_row=16*i+10, end_column=4)
        sheet.merge_cells(start_row=16*i+11, start_column=4, end_row=16*i+14, end_column=4)
        sheet.merge_cells(start_row=16*i+15, start_column=4, end_row=16*i+18, end_column=4)
    sheet.merge_cells(start_row=83, start_column=2, end_row=83, end_column=4)
    sheet.merge_cells(start_row=84, start_column=2, end_row=84, end_column=4)
    sheet.merge_cells(start_row=85, start_column=2, end_row=85, end_column=4)

    # 设置单元格字体
    font = Font(name=u'宋体', size=9, bold=True)
    font2 = Font(name=u'宋体', size=9)
    sheet.cell(row=1, column=i+1).font = font
    # 格式
    align = Alignment(horizontal='center', vertical='center')
    fill = PatternFill(start_color='CC99CC', end_color='CC99CC', fill_type='solid')
    fill2 = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    sheet.freeze_panes = 'F3'
    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000'))

    for i in range(2,86):
        for j in range(2,37):
            sheet.cell(row=i, column=j).border = border

    for i in range(0,20):
        for j in range(5,37):
            sheet.cell(row=4*i+6, column=j).fill = fill2

    # 3、填写内容
    for i in range(len(context_list)):
        sheet.cell(row=2, column=i+2).alignment = align
        sheet.cell(row=2, column=i+2, value=10).fill = fill
        sheet.cell(row=2, column=i+2).font = font
        sheet.cell(row=2, column=i+2).value = context_list[i]

    B_list = ['一', '二', '三', '四', '五']
    sheet.cell(row=3, column=2).alignment = align
    sheet.cell(row=3, column=2).font = font
    sheet.cell(row=3, column=2).value = B_list[0]
    sheet.cell(row=19, column=2).alignment = align
    sheet.cell(row=19, column=2).font = font
    sheet.cell(row=19, column=2).value = B_list[1]
    sheet.cell(row=35, column=2).alignment = align
    sheet.cell(row=35, column=2).font = font
    sheet.cell(row=35, column=2).value = B_list[2]
    sheet.cell(row=51, column=2).alignment = align
    sheet.cell(row=51, column=2).font = font
    sheet.cell(row=51, column=2).value = B_list[3]
    sheet.cell(row=67, column=2).alignment = align
    sheet.cell(row=67, column=2).font = font
    sheet.cell(row=67, column=2).value = B_list[4]

    C_list = ['一个月', '两个月', '三个月', '四个月']
    for j in range(len(C_list)):
        sheet.cell(row=3+4*j, column=3).alignment = align
        sheet.cell(row=3+4*j, column=3).font = font
        sheet.cell(row=3+4*j, column=3).value = C_list[j]
        sheet.cell(row=19+4*j, column=3).alignment = align
        sheet.cell(row=19+4*j, column=3).font = font
        sheet.cell(row=19+4*j, column=3).value = C_list[j]
        sheet.cell(row=35+4*j, column=3).alignment = align
        sheet.cell(row=35+4*j, column=3).font = font
        sheet.cell(row=35+4*j, column=3).value = C_list[j]
        sheet.cell(row=51+4*j, column=3).alignment = align
        sheet.cell(row=51+4*j, column=3).font = font
        sheet.cell(row=51+4*j, column=3).value = C_list[j]
        sheet.cell(row=67+4*j, column=3).alignment = align
        sheet.cell(row=67+4*j, column=3).font = font
        sheet.cell(row=67+4*j, column=3).value = C_list[j]

    D_list = ["1签约","2签约","3交付","/","1签约",
              "2签约","3交付","/","1签约","2签约",
              "3交付","/","1签约","2签约","3交付",
              "/","1签约","2签约","3交付","/"]
    for i in range(len(D_list)):
        sheet.cell(row=4*i+3, column=4).alignment = align
        sheet.cell(row=4 * i + 3, column=4).font = font
        sheet.cell(row=4*i+3, column=4).value = D_list[i]

    E_list=['叶1','叶2','叶3','叶4']
    for t in range(len(E_list)):
        for l in range(0,20):
            sheet.cell(row=t + (4*l+3), column=5).alignment = align
            sheet.cell(row=t + (4 * l + 3), column=5).font = font2
            sheet.cell(row=t + (4*l+3), column=5).value = E_list[t]

    row_list = ["成绩","分数","百分比"]
    sheet.cell(row=83, column=2).alignment = align
    sheet.cell(row=83, column=2).font = font
    sheet.cell(row=83, column=2).value = row_list[0]
    sheet.cell(row=84, column=2).alignment = align
    sheet.cell(row=84, column=2).font = font
    sheet.cell(row=84, column=2).value = row_list[1]
    sheet.cell(row=85, column=2).alignment = align
    sheet.cell(row=85, column=2).font = font
    sheet.cell(row=85, column=2).value = row_list[2]

    # 4、保存文件
    curPath = os.getcwd()
    tempPath = 'file_target'
    targetPath = curPath + os.path.sep + tempPath
    if not os.path.exists(targetPath):
        os.makedirs(targetPath)
    for i in range(len(title_list)):
        f.save(targetPath+"/"+title_list[i]+".xlsx")

if __name__ == '__main__':
    context_list = ["季度","期别","对象","案例","合计","项目1","项目2",
                    "项目3","项目4","项目5","项目6","项目7","项目8","项目9",
                    "项目10","项目11","项目12","项目13","项目14","项目15",
                    "项目16","项目17","项目18","项目19","项目20","项目21",
                    "项目22","项目23","项目24","项目25","项目26","项目27",
                    "项目28","项目29","项目30",]
    title_list = ['2018XXXXXX']
    sheet_list = ['4','说明','合计','案例','sheet3']
    create_excel(context_list,title_list,sheet_list)



























