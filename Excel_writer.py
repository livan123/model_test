#!/usr/bin/env python
# _*_ UTF-8 _*_
'''
@project:GBDT+LR-Demo
@author:xiaofei
'''
import openpyxl
# 写excel

# 用来存储数据
tables = []
newTables = []
def write_excel():
    f = openpyxl.Workbook()  # 创建工作簿
    sheet1 = f.create_sheet()
    # sheet1 = f.add_sheet(u'sheet1',cell_overwrite_ok=True) #创建sheet
    row0 = [u'L1', u'L2', u'L3', u'L4', u'问题', u'答案']
    # 生成第一行
    # for i in range(len(row0)):
    #    sheet1.cell(column=i,row=0).value='L1')
    # 生成后续
    for jkey in range(len(newTables)):
        jk = 1
        for cT in range(arrayNum):
            jk = jkey + 1
            if cT == 0:
                sheet1.cell(row=jk, column=cT + 1).value = '1'
            else:
                sheet1.cell(row=jk, column=cT + 1).value = '2'
    f.save("chatPy.xlsx")  # 保存文件
if __name__ == '__main__':
    # 写入Excel
    write_excel();
    print('写入成功')
